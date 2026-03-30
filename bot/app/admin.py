from maxapi import Router
from maxapi.types import (
    MessageCreated,
    Command,
    CommandStart,
    Message,
    MessageCallback,
    BotStarted,
)
from database.Database import DataBase
from database.models import QuestionnaireStatus
from database.models import *
from core.log import Logger
from core.dictionary import *
from maxapi.context import MemoryContext
import app.keyboards as kb
from app.scheduler import BotScheduler
from maxapi import F
from maxapi.filters.callback_payload import CallbackPayload
from typing import Any, Dict
from datetime import datetime
from app.states import *
import asyncio
from maxapi.enums.parse_mode import ParseMode
import re
from datetime import datetime, date, time
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from maxapi.types import InputMedia
import io
import os
import tempfile
from maxapi.utils.inline_keyboard import InlineKeyboardBuilder
from maxapi.types import CallbackButton

router = Router()
logger = Logger(__name__)


######################################################################
@router.message_created(Command("id", prefix="#"))
async def handle_id_command(event: MessageCreated):
    await event.message.answer(f"Ваш ID: {event.chat.chat_id}")

#####################################################################
@router.message_callback(F.callback.payload == 'admin_help')
async def admin_help(event: MessageCallback, db: DataBase):
    """Справка для администратора"""
    user_id = event.chat.chat_id
    user_state = await db.get_state(user_id)
    
    # Удаляем предыдущие сообщения
    if user_state and user_state.last_message_ids:
        await db.delete_messages(user_state)
        user_state.last_message_ids = []
        await db.update_state(user_state)
    
    if not await db.check_admin(user_id):
        await event.message.answer("⛔ У вас нет доступа", attachments=[kb.home_keyboard_inline()])
        return
    
    help_text = (
        "🔐 <b>Административная справка</b>\n\n"
        
        "📌 <b>Команды:</b>\n"
        "• /admin - открыть админ-панель\n"
        "• #id - узнать свой ID\n\n"
        
        "👑 <b>Управление администраторами:</b>\n"
        "• #addadmin ID - добавить админа\n"
        "• #deladmin ID - удалить админа\n"
        "• #admins - список админов\n\n"
        
        "📞 <b>Управление регистраторами:</b>\n"
        "• #addstaff ID - добавить регистратора\n"
        "• #delstaff ID - удалить регистратора\n"
        "• #staff - список регистраторов\n\n"
        
        "🛠️ <b>Админ-панель (/admin):</b>\n"
        "• Управление врачами\n"
        "• Настройка расписания\n"
        "• Статистика и экспорт"
    )
    
    sent_msg = await event.message.answer(
        help_text,
        attachments=[kb.admin_kb()]
    )
    
    # Сохраняем ID нового сообщения
    user_state.last_message_ids.append(sent_msg.message.body.mid)
    await db.update_state(user_state)
#####################################################################
@router.message_created(Command("addadmin", prefix="#"))
async def add_admin(event: MessageCreated, db: DataBase):
    """Добавить администратора: #addadmin ID"""
    # Только текущий админ может добавлять
    if not await db.check_admin(event.chat.chat_id):
        await event.message.answer("⛔ У вас нет прав")
        return
    
    args = event.message.body.text.split()
    if len(args) < 2:
        await event.message.answer("❌ Использование: #addadmin ID")
        return
    
    try:
        user_id = int(args[1])
        username = args[2] if len(args) > 2 else None
        
        if await db.add_staff(user_id, role=0, username=username):
            await event.message.answer(f"✅ Администратор {user_id} добавлен")
        else:
            await event.message.answer(f"❌ Ошибка: пользователь {user_id} уже администратор")
    except ValueError:
        await event.message.answer("❌ ID должен быть числом")

##################################################################
@router.message_created(Command("deladmin", prefix="#"))
async def del_admin(event: MessageCreated, db: DataBase):
    """Удалить администратора: #deladmin ID"""
    if not await db.check_admin(event.chat.chat_id):
        await event.message.answer("⛔ У вас нет прав")
        return
    
    args = event.message.body.text.split()
    if len(args) < 2:
        await event.message.answer("❌ Использование: #deladmin ID")
        return
    
    try:
        user_id = int(args[1])
        if await db.remove_staff(user_id, role=0):
            await event.message.answer(f"✅ Администратор {user_id} удален")
        else:
            await event.message.answer(f"❌ Администратор {user_id} не найден")
    except ValueError:
        await event.message.answer("❌ ID должен быть числом")

#########################################################
@router.message_created(Command("addstaff", prefix="#"))
async def add_registration_staff(event: MessageCreated, db: DataBase):
    """Добавить регистратора: #addstaff ID"""
    if not await db.check_admin(event.chat.chat_id):
        await event.message.answer("⛔ У вас нет прав")
        return
    
    args = event.message.body.text.split()
    if len(args) < 2:
        await event.message.answer("❌ Использование: #addstaff ID")
        return
    
    try:
        user_id = int(args[1])
        username = args[2] if len(args) > 2 else None
        
        if await db.add_staff(user_id, role=1, username=username):
            await event.message.answer(f"✅ Регистратор {user_id} добавлен")
        else:
            await event.message.answer(f"❌ Ошибка: пользователь {user_id} уже регистратор")
    except ValueError:
        await event.message.answer("❌ ID должен быть числом")

###########################################################################
@router.message_created(Command("delstaff", prefix="#"))
async def del_registration_staff(event: MessageCreated, db: DataBase):
    """Удалить регистратора: #delstaff ID"""
    if not await db.check_admin(event.chat.chat_id):
        await event.message.answer("⛔ У вас нет прав")
        return
    
    args = event.message.body.text.split()
    if len(args) < 2:
        await event.message.answer("❌ Использование: #delstaff ID")
        return
    
    try:
        user_id = int(args[1])
        if await db.remove_staff(user_id, role=1):
            await event.message.answer(f"✅ Регистратор {user_id} удален")
        else:
            await event.message.answer(f"❌ Регистратор {user_id} не найден")
    except ValueError:
        await event.message.answer("❌ ID должен быть числом")

###################################################################
@router.message_created(Command("admins", prefix="#"))
async def list_admins(event: MessageCreated, db: DataBase):
    """Список администраторов"""
    if not await db.check_admin(event.chat.chat_id):
        await event.message.answer("⛔ У вас нет прав")
        return
    
    admins = await db.get_staff_by_role(0)
    if not admins:
        await event.message.answer("📭 Список администраторов пуст")
        return
    
    text = "👥 <b>Администраторы:</b>\n\n"
    for admin in admins:
        text += f"• ID: {admin.id_max}"
        if admin.username:
            text += f" (@{admin.username})"
        text += "\n"
    
    await event.message.answer(text)

###############################################################
@router.message_created(Command("staff", prefix="#"))
async def list_registration_staff(event: MessageCreated, db: DataBase):
    """Список регистраторов"""
    if not await db.check_admin(event.chat.chat_id):
        await event.message.answer("⛔ У вас нет прав")
        return
    
    staff_list = await db.get_staff_by_role(1)
    if not staff_list:
        await event.message.answer("📭 Список регистраторов пуст")
        return
    
    text = "📞 <b>Регистраторы:</b>\n\n"
    for staff in staff_list:
        text += f"• ID: {staff.id_max}"
        if staff.username:
            text += f" (@{staff.username})"
        text += "\n"
    
    await event.message.answer(text)

#####################################################################
@router.message_callback(F.callback.payload == "admin")
@router.message_created(Command("admin"))
@router.message_created(F.text == CANCEL_ENTER)
async def admin_panel(
    event: MessageCreated | MessageCallback, context: MemoryContext, db: DataBase
):
    """Обработчик админ панель (работает и для сообщений и для callback)"""

    try:
        # Очищаем состояние
        await context.set_state(None)

        # Определяем тип события и получаем необходимые объекты
        if isinstance(event, MessageCallback):
            message = event.message
            user = event.chat.chat_id
        else:
            message = event
            user = event.chat.chat_id

        # Получаем или создаем состояние пользователя
        user_state = await db.get_state(event.chat.chat_id)
        # if not user_state:
        #     user_state = await db.create_state(user.id)  # Предполагаем, что такой метод есть

        # Удаляем предыдущие сообщения если нужно
        if user_state.last_message_ids:
            await db.delete_messages(user_state)
            user_state.last_message_ids = []

        await logger.info(f"ID_MAX:{event.chat.chat_id}|Команда admin")

        # Проверяем права администратора
        if not await db.check_admin(event.chat.chat_id):
            await logger.warning(
                f"Unauthorized admin access attempt by {event.chat.chat_id}"
            )
            sent_msg = await event.message.answer(
                "⛔️ У вас нет доступа к админ-панели",
                attachments=[kb.home_keyboard_inline()],
            )
            user_state.last_message_ids.append(sent_msg.message.body.mid)
            await db.update_state(user_state)
            return

        # Получаем текущие настройки бота
        settings = await db.get_settings()

        # Отправляем админ-панель
        sent_msg = await message.answer(
            "🛠️ Административная панель",
            attachments=[
                kb.admin_main_kb(
                    bot_status=settings.bot_active, auto_schedule=settings.auto_schedule
                )
            ],
        )

        # Обновляем состояние пользователя
        user_state.last_message_ids.append(sent_msg.message.body.mid)
        await db.update_state(user_state)

    except Exception as e:
        await logger.error(f"Ошибка входа в админ-панель: {e}")
        error_msg = "⚠️ Произошла ошибка. Попробуйте позже."
        if isinstance(event, MessageCallback):
            await event.message.answer(
                error_msg, attachments=[kb.home_keyboard_inline()]
            )
        else:
            await event.message.answer(
                error_msg, attachments=[kb.home_keyboard_inline()]
            )
    finally:
        await db.close()


#####################################################################
@router.message_callback(F.callback.payload.startswith("toggle_bot_"))
async def toggle_bot_handler(
    callback: MessageCallback, db: DataBase, scheduler: "BotScheduler"
):
    await logger.info(f"ID_MAX:{callback.chat.chat_id}|=>toggle_bot_handler")
    try:
        # Получаем текущие настройки
        settings = await db.get_settings()
        new_status = not settings.bot_active

        # Обновляем настройки в БД
        await db.update_settings(
            user_id=callback.chat.chat_id,
            bot_active=new_status,
            manual_override=True,
            last_changed=datetime.now(),  # Важно для отслеживания изменений
        )

        # 1. Сначала отвечаем на callback
        await callback.answer()

        # 2. Полностью пересоздаем сообщение с новой клавиатурой

        await callback.message.edit(
            text="🛠️ Административная панель",
            attachments=[kb.admin_main_kb(new_status, settings.auto_schedule)],
        )

        # При включении проверяем расписание
        if new_status:
            await scheduler._check_schedule()

    except Exception as e:
        await logger.error(f"Ошибка переключения бота: {e}")
        await callback.message.answer(
            "⚠️ Ошибка изменения статуса", attachments=[kb.home_keyboard_inline()]
        )


#####################################################################


#####################################################################
@router.message_callback(F.callback.payload == "toggle_auto_schedule")
async def toggle_auto_schedule_handler(
    callback: MessageCallback,
    db: DataBase,
    scheduler: "BotScheduler",  # Используем строковую аннотацию для избежания циклических импортов
):
    """
    Обработчик переключения режима авторасписания
    Args:
        callback: CallbackQuery объект
        db: Объект работы с БД
        scheduler: Объект планировщика бота
    """
    try:
        # Получаем текущие настройки
        settings = await db.get_settings()

        # Инвертируем текущее состояние авторасписания
        new_auto_schedule = not settings.auto_schedule

        # Обновляем настройки в БД
        await db.update_settings(
            user_id=callback.chat.chat_id,
            auto_schedule=new_auto_schedule,
            manual_override=False if new_auto_schedule else settings.manual_override,
        )

        # Если включаем авторасписание - принудительно проверяем текущее расписание
        if new_auto_schedule:
            await scheduler._check_schedule()

        # Обновляем сообщение с админ-панелью
        # await update_admin_message(callback, db)
        settings = await db.get_settings()
        await callback.message.edit(
            text="🛠️ Административная панель",
            attachments=[kb.admin_main_kb(settings.bot_active, settings.auto_schedule)],
        )
        # Отправляем уведомление
        # status = "включено" if new_auto_schedule else "выключено"
        # await callback.answer(f"Авторасписание {status}")

    except Exception as e:
        await logger.error(f"Ошибка при переключении авторасписания: {e}")
        await callback.message.answer(
            "⚠️ Ошибка изменения Авторасписания", attachments=[kb.home_keyboard_inline()]
        )


#####################################################################
@router.message_callback(F.callback.payload == "edit_schedule")
async def edit_schedule_handler(callback: MessageCallback, db: DataBase):
    """Обработчик настройки часов работы"""
    try:
        user_state = await db.get_state(callback.chat.chat_id)
        if user_state.last_message_ids:
            await db.delete_messages(user_state)
            user_state.last_message_ids = []
        sent_mess = await callback.message.answer(
            "⏰ Настройка часов работы:",
            attachments=[await kb.schedule_settings_kb(db)],
        )
        user_state.last_message_ids.append(sent_mess.message.body.mid)
        await db.update_state(user_state)
    except Exception as e:
        logger.error(f"Ошибка настройки расписания: {e}")
        await callback.message.answer(
            "⚠️ Ошибка открытия настроек", attachments=[kb.home_keyboard_inline()]
        )


#####################################################################
@router.message_callback(F.callback.payload == "weekday_time_setup")
async def handle_weekday_time_setup(
    event: MessageCallback, db: DataBase, context: MemoryContext
):
    """Обработчик изменения времени по будням"""
    # Очистка предыдущих сообщений если нужно
    user_state = await db.get_state(event.chat.chat_id)
    if user_state and user_state.last_message_ids:
        await db.delete_messages(user_state)

    await logger.info(f"ID_MAX:{event.chat.chat_id}|=>handle_weekday_time_setup")

    # Устанавливаем состояние ожидания времени начала
    await context.set_state(TimeSetupStates.WAITING_WEEKDAY_START)

    # Сохраняем ID сообщения для последующего удаления
    sent_mess = await event.message.answer(
        "🕘 Введите время НАЧАЛА работы в будни в формате ЧЧ:ММ (например, 08:00):"
    )
    user_state.last_message_ids.append(sent_mess.message.body.mid)
    await db.update_state(user_state)


#####################################################################
@router.message_created(TimeSetupStates.WAITING_WEEKDAY_START)
async def process_weekday_start_time(
    event: MessageCreated, db: DataBase, context: MemoryContext
):
    """Обработка времени начала работы в будни"""
    # Очистка предыдущих сообщений если нужно
    user_state = await db.get_state(event.chat.chat_id)
    if user_state and user_state.last_message_ids:
        await db.delete_messages(user_state)

    await logger.info(f"ID_MAX:{event.chat.chat_id}|=>process_weekday_start_time")
    if not re.match(r"^([0-1]?[0-9]|2[0-3]):[0-5][0-9]$", event.message.body.text):
        msg = await event.message.answer(
            "❌ Неверный формат времени. Используйте ЧЧ:ММ (например, 08:00)"
        )
        user_state.last_message_ids.append(msg.message.body.mid)
        await db.update_state(user_state)
        return

    # Сохраняем время начала
    await context.update_data(weekday_start=event.message.body.text)

    # Переходим к ожиданию времени окончания
    await context.set_state(TimeSetupStates.WAITING_WEEKDAY_END)

    msg = await event.message.answer(
        f"⏱ Введите время ОКОНЧАНИЯ работы в будни (текущее начало: {event.message.body.text}):"
    )
    user_state.last_message_ids.append(msg.message.body.mid)
    await db.update_state(user_state)

    # Удаляем сообщение пользователя с временем
    await event.message.delete()


#####################################################################
@router.message_created(TimeSetupStates.WAITING_WEEKDAY_END)
async def process_weekday_end_time(
    event: MessageCreated, db: DataBase, context: MemoryContext
):
    """Обработка времени окончания работы в будни"""
    # Очистка предыдущих сообщений если нужно
    user_state = await db.get_state(event.chat.chat_id)
    if user_state and user_state.last_message_ids:
        await db.delete_messages(user_state)

    await logger.info(f"ID_MAX:{event.chat.chat_id}|=>process_weekday_end_time")

    if not re.match(r"^([0-1]?[0-9]|2[0-3]):[0-5][0-9]$", event.message.body.text):
        msg = await event.message.answer(
            "❌ Неверный формат времени. Используйте ЧЧ:ММ (например, 17:00)"
        )
        user_state.last_message_ids.append(msg.message.body.mid)
        await db.update_state(user_state)
        return

    # Получаем сохраненное время начала
    data = await context.get_data()
    start_time_str = data.get("weekday_start")
    end_time_str = event.message.body.text

    # Преобразуем в объекты time
    try:
        start_time = time(*map(int, start_time_str.split(":")))
        end_time = time(*map(int, end_time_str.split(":")))

        # Проверяем что время окончания позже начала
        if end_time <= start_time:
            msg = await event.message.answer(
                "❌ Время окончания должно быть позже времени начала"
            )
            user_state.last_message_ids.append(msg.message.body.mid)
            await db.update_state(user_state)
            return

        # Обновляем настройки в БД
        await db.update_settings(
            user_id=event.chat.chat_id,
            weekday_start=start_time,
            weekday_end=end_time,
            last_changed=datetime.now(),
        )

        # Возвращаемся в меню настроек
        await context.set_state(None)

        # Удаляем сообщение пользователя
        await event.message.delete()

        # Показываем обновленные настройки
        keyboard = [await kb.schedule_settings_kb(db)]
        msg = await event.message.answer(
            f"✅ Часы работы в будни обновлены: {start_time_str}-{end_time_str}",
            attachments=keyboard,
        )
        user_state.last_message_ids.append(msg.message.body.mid)
        await db.update_state(user_state)

    except Exception as e:
        await logger.error(f"Ошибка обновления времени будней: {e}")
        msg = await event.message.answer(
            "❌ Произошла ошибка при обновлении расписания",
            attachments=[kb.home_keyboard_inline()],
        )
        user_state.last_message_ids.append(msg.message.body.mid)
        await db.update_state(user_state)


#####################################################################
# Обработчик запуска настройки выходных
@router.message_callback(F.callback.payload == "weekend_time_setup")
async def handle_weekend_time_setup(event: MessageCallback, db: DataBase, context: MemoryContext):
    """Обработчик изменения времени по выходным"""
    user_state = await db.get_state(event.chat.chat_id)
    if user_state and user_state.last_message_ids:
        await db.delete_messages(user_state)

    await logger.info(f'ID_MAX:{event.chat.chat_id}|=>handle_weekend_time_setup')
    
    await context.set_state(TimeSetupStates.WAITING_WEEKEND_START)
    
    sent_mess = await event.message.answer(
        "🌅 Введите время НАЧАЛА работы в выходные в формате ЧЧ:ММ (например, 09:00):"
    )
    user_state.last_message_ids.append(sent_mess.message.body.mid)
    await db.update_state(user_state)
    
###############################################################    
# Обработка времени начала работы в выходные
@router.message_created(TimeSetupStates.WAITING_WEEKEND_START)
async def process_weekend_start_time(event: MessageCreated, db: DataBase, context: MemoryContext):
    """Обработка времени начала работы в выходные"""
    user_state = await db.get_state(event.chat.chat_id)
    if user_state and user_state.last_message_ids:
        await db.delete_messages(user_state)

    await logger.info(f'ID_TG:{event.chat.chat_id}|=>process_weekend_start_time')
    
    if not re.match(r'^([0-1]?[0-9]|2[0-3]):[0-5][0-9]$', event.message.body.text):
        msg = await event.message.answer("❌ Неверный формат времени. Используйте ЧЧ:ММ (например, 09:00)")
        user_state.last_message_ids.append(msg.message.body.mid)
        await db.update_state(user_state)
        return
    
    await context.update_data(weekend_start=event.message.body.text)
    await context.set_state(TimeSetupStates.WAITING_WEEKEND_END)
    
    msg = await event.message.answer(
        f"⏱ Введите время ОКОНЧАНИЯ работы в выходные (текущее начало: {event.message.body.text}):"
    )
    user_state.last_message_ids.append(msg.message.body.mid)
    await db.update_state(user_state)
    await event.message.delete()
############################################################
# Обработка времени окончания работы в выходные
@router.message_created(TimeSetupStates.WAITING_WEEKEND_END)
async def process_weekend_end_time(event: MessageCreated, db: DataBase, context: MemoryContext):
    """Обработка времени окончания работы в выходные"""
    user_state = await db.get_state(event.chat.chat_id)
    if user_state and user_state.last_message_ids:
        await db.delete_messages(user_state)

    await logger.info(f'ID_MAX:{event.chat.chat_id}|=>process_weekend_end_time')

    if not re.match(r'^([0-1]?[0-9]|2[0-3]):[0-5][0-9]$', event.message.body.text):
        msg = await event.message.answer("❌ Неверный формат времени. Используйте ЧЧ:ММ (например, 14:00)")
        user_state.last_message_ids.append(msg.message.body.mid)
        await db.update_state(user_state)
        return
    
    data = await context.get_data()
    start_time_str = data.get('weekend_start')
    end_time_str = event.message.body.text
    
    try:
        start_time = time(*map(int, start_time_str.split(':')))
        end_time = time(*map(int, end_time_str.split(':')))
        
        if end_time <= start_time:
            msg = await event.message.answer("❌ Время окончания должно быть позже времени начала")
            user_state.last_message_ids.append(msg.message.body.mid)
            await db.update_state(user_state)
            return
            
        await db.update_settings(
            user_id=event.chat.chat_id,
            weekend_start=start_time,
            weekend_end=end_time,
            last_changed=datetime.now()
        )
        
        await context.set_state(None)
        await event.message.delete()
        
        keyboard = [await kb.schedule_settings_kb(db)]
        msg = await event.message.answer(
            f"✅ Часы работы в выходные обновлены: {start_time_str}-{end_time_str}",
            attachments=keyboard
        )
        user_state.last_message_ids.append(msg.message.body.mid)
        await db.update_state(user_state)
        
    except Exception as e:
        await logger.error(f"Ошибка обновления времени выходных: {e}")
        msg = await event.message.answer("❌ Произошла ошибка при обновлении расписания", attachments=[kb.home_keyboard_inline()])
        user_state.last_message_ids.append(msg.message.body.mid)
        await db.update_state(user_state)
################################################################   
@router.message_callback(F.callback.payload.startswith("sotrudniki"))
async def handle_sotrudniki(event: MessageCallback, db: DataBase, context: MemoryContext):
    """Обработчик сотрудники"""
    await logger.info("sotrudniki")
    user_id = event.chat.chat_id
    user_state = await db.get_state(user_id)
    # if user_state.last_message_ids:
    #         await db.delete_messages(user_state)
    #         user_state.last_message_ids = []
    # Загружаем список сотрудников
    doctor_lst = await db.get_all_doctors()   
    await logger.info(f"doctor_lst:{doctor_lst}")

    if not doctor_lst:
        sent_mess = await event.message.answer(
            "👨‍⚕️ Список врачей пуст",
            attachments=[await kb.inline_add_doctor()]
        )
        user_state.last_message_ids.append(sent_mess.message.body.mid)
        await db.update_state(user_state)
        return

    # Получаем текущую позицию из состояния
    state_data = await context.get_data()
    current_position = state_data.get('position', 0)
    
    # Обработка перехода по пагинации
    payload = event.callback.payload
    
    if payload.startswith("sotrudniki_pages"):
        parts = payload.split(":")
        if len(parts) >= 3:
            try:
                new_position = int(parts[1])
                current_position = new_position
                await logger.info(f"Переход на позицию: {current_position}")
            except ValueError as e:
                await logger.error(f"Ошибка парсинга позиции: {e}")
    
    # Сохраняем новую позицию в состоянии
    await context.update_data(position=current_position)

    try:
        total_items = len(doctor_lst)
        total_pages = total_items
        
        # Проверяем границы
        if current_position >= total_items:
            current_position = total_items - 1
            await context.update_data(position=current_position)
        if current_position < 0:
            current_position = 0
            await context.update_data(position=current_position)
        
        current_doctor = doctor_lst[current_position]
        
        # Форматирование информации о докторе
        def _format_doctor_info(doctor):
            created_at_str = doctor.created_at.strftime('%d.%m.%Y') if doctor.created_at else "неизвестно"
            phone = getattr(doctor, 'phone', 'не указан')
            if not phone:
                phone = "не указан"
            
            return (
                f"👨‍⚕️ <b>Врач {current_position + 1} из {total_items}</b>\n\n"
                f"👤 <b>ФИО:</b> {doctor.full_name}\n"
                f"📞 <b>Телефон:</b> {phone}\n"
                f"🆔 <b>ID:</b> {doctor.id}\n"
                f"🕒 <b>Активен:</b> {'✅ Да' if doctor.is_active else '❌ Нет'}\n"
                f"📅 <b>Зарегистрирован:</b> {created_at_str}"
            )

        # Создаем клавиатуру
        pagination_kb = await kb.inline_pagination(
            position=current_position,
            pages=total_pages,
            id=current_doctor.id
        )

        # Редактируем текущее сообщение
        await event.message.edit(
            text=_format_doctor_info(current_doctor),
            attachments=[pagination_kb]
        )
        await logger.info(f"Отредактировано сообщение")
        
        # Сохраняем ID сообщения в состоянии, если нужно
        await context.update_data(sent_mess_id=event.message.body.mid)
        
    except Exception as e:
        await logger.error(f"Ошибка: {e}")
        import traceback
        traceback.print_exc()     

################################################################
@router.message_callback(F.callback.payload.startswith("change_status"))
async def handle_change_status(event: MessageCallback, db: DataBase, context: MemoryContext):
    """Обработчик изменения статуса доктора (активен/неактивен)"""
    try:
        # Получаем ID доктора из callback данных
        _, id_doctor = event.callback.payload.split(":")
        id_doctor = int(id_doctor)
        
        # Меняем статус доктора
        success = await db.change_status_doctor(id_doctor)
        
        if not success:
            await event.message.answer("❌ Ошибка при изменении статуса")
            return
        
        await logger.info(f"Статус доктора {id_doctor} изменен")
        
        # Получаем обновленного доктора
        doctor = await db.get_doctor(id_doctor)
        if not doctor:
            await event.message.answer("❌ Доктор не найден")
            return
        
        # Получаем список всех врачей для пагинации
        doctor_lst = await db.get_all_doctors()
        
        # Получаем текущую позицию из состояния
        state_data = await context.get_data()
        current_position = state_data.get('position', 0)
        
        # Находим позицию текущего доктора в списке
        for i, d in enumerate(doctor_lst):
            if d.id == id_doctor:
                current_position = i
                break
        
        total_items = len(doctor_lst)
        total_pages = total_items
        
        # Форматирование информации о докторе с новым статусом
        def _format_doctor_info(doctor, position, total):
            created_at_str = doctor.created_at.strftime('%d.%m.%Y') if doctor.created_at else "неизвестно"
            phone = getattr(doctor, 'phone', 'не указан')
            if not phone:
                phone = "не указан"
            status_text = "✅ Да" if doctor.is_active else "❌ Нет"
            
            return (
                f"👨‍⚕️ <b>Врач {position + 1} из {total}</b>\n\n"
                f"👤 <b>ФИО:</b> {doctor.full_name}\n"
                f"📞 <b>Телефон:</b> {phone}\n"
                f"🆔 <b>ID:</b> {doctor.id}\n"
                f"🕒 <b>Активен:</b> {status_text}\n"
                f"📅 <b>Зарегистрирован:</b> {created_at_str}"
            )
        
        # Создаем клавиатуру с обновленным статусом
        pagination_kb = await kb.inline_pagination(
            position=current_position,
            pages=total_pages,
            id=doctor.id
        )
        
        # Редактируем текущее сообщение
        await event.message.edit(
            text=_format_doctor_info(doctor, current_position, total_items),
            attachments=[pagination_kb]
        )
        
        await logger.info(f"Сообщение отредактировано с новым статусом доктора {doctor.full_name}")
        
    except ValueError:
        await logger.error("Некорректный ID доктора")
        await event.message.answer("Некорректный ID доктора")
    except Exception as e:
        await logger.error(f"Ошибка в handle_change_status: {e}")
        import traceback
        traceback.print_exc()
        await event.message.answer("Произошла ошибка")
###############################################################
@router.message_callback(F.callback.payload.startswith("delete_doctor"))
async def handle_delete_doctor(event: MessageCallback, db: DataBase, context: MemoryContext):
    """Обработчик удаления доктора"""
    try:
        # Получаем ID доктора из callback данных
        _, id_doctor = event.callback.payload.split(":")
        id_doctor = int(id_doctor)
        
        # Получаем текущую позицию из состояния
        state_data = await context.get_data()
        current_position = state_data.get('position', 0)
        
        # Удаляем доктора
        success_del = await db.doctor_delete(id_doctor)
        
        if not success_del:
            await event.message.answer("❌ Ошибка при удалении доктора")
            return
        
        await logger.info(f"Доктор {id_doctor} удален")
        
        # Получаем обновленный список врачей
        doctor_lst = await db.get_all_doctors()
        total_items = len(doctor_lst)
        
        # Если список врачей пуст
        if total_items == 0:
            # Показываем сообщение, что врачей нет
            await event.message.edit(
                text="👨‍⚕️ Список врачей пуст\n\nДобавьте нового врача с помощью кнопки ниже.",
                attachments=[await kb.inline_add_doctor()]
            )
            # Очищаем состояние
            await context.update_data(sent_mess_id=None, position=0)
            return
        
        # Корректируем позицию (если удалили последнего, показываем предыдущего)
        if current_position >= total_items:
            current_position = total_items - 1
        
        # Получаем текущего доктора
        current_doctor = doctor_lst[current_position]
        total_pages = total_items
        
        # Форматирование информации о докторе
        def _format_doctor_info(doctor, position, total):
            created_at_str = doctor.created_at.strftime('%d.%m.%Y') if doctor.created_at else "неизвестно"
            phone = getattr(doctor, 'phone', 'не указан')
            if not phone:
                phone = "не указан"
            status_text = "✅ Да" if doctor.is_active else "❌ Нет"
            
            return (
                f"👨‍⚕️ <b>Врач {position + 1} из {total}</b>\n\n"
                f"👤 <b>ФИО:</b> {doctor.full_name}\n"
                f"📞 <b>Телефон:</b> {phone}\n"
                f"🆔 <b>ID:</b> {doctor.id}\n"
                f"🕒 <b>Активен:</b> {status_text}\n"
                f"📅 <b>Зарегистрирован:</b> {created_at_str}"
            )
        
        # Создаем клавиатуру
        pagination_kb = await kb.inline_pagination(
            position=current_position,
            pages=total_pages,
            id=current_doctor.id
        )
        
        # Сохраняем новую позицию в состоянии
        await context.update_data(position=current_position)
        
        # Редактируем текущее сообщение с новым доктором
        await event.message.edit(
            text=_format_doctor_info(current_doctor, current_position, total_items),
            attachments=[pagination_kb]
        )
        
        await logger.info(f"Сообщение отредактировано, показываем доктора {current_doctor.full_name}")
        
    except ValueError:
        await logger.error("Некорректный ID доктора")
        await event.message.answer("Некорректный ID доктора")
    except Exception as e:
        await logger.error(f"Ошибка в handle_delete_doctor: {e}")
        import traceback
        traceback.print_exc()
        await event.message.answer("Произошла ошибка")

#############################################################
# add_doctor
@router.message_callback(F.callback.payload == "add_doctor")
async def handle_add_doctor(event: MessageCallback, db: DataBase, context: MemoryContext):
    """Обработчик добавления доктора"""
    user_id = event.chat.chat_id
    user_state = await db.get_state(user_id)
    
    # Удаляем предыдущие сообщения если нужно
    if user_state and user_state.last_message_ids:
        await db.delete_messages(user_state)
    try:
        sent_mess = await event.message.answer(
                "1. 📱 Укажите <b>ID MAX</b><i>(команда #id)</i>:"
                )
        
        # Не сохраняем ID этого сообщения для последующего удаления
        await context.set_state(AddDoctor.max_id)
        # Обновляем состояние пользователя
        user_state.last_message_ids.append(sent_mess.message.body.mid)
        await db.update_state(user_state)  


    except Exception as e:
        await logger.error(f"Ошибка начала анкеты: {e}")
        await event.message.answer("⚠️ Ошибка. Попробуйте позже.", attachments=[kb.home_keyboard_inline()]) 

##########################################################################        
@router.message_created(AddDoctor.max_id)
async def process_add_tg_id(event: MessageCreated, db: DataBase, context: MemoryContext):
    """Обработчик для ввода Telegram ID доктора"""
    try:
        # Проверяем что введено число
        max_id = int(event.message.body.text.strip())
        
        # Дополнительные проверки
        if max_id <= 0:
            await event.message.answer("❌ ID должен быть положительным числом. Попробуйте еще раз:")
            return
            
        if len(str(max_id)) < 5:
            await event.message.answer("❌ ID слишком короткий. Минимум 5 цифр. Попробуйте еще раз:")
            return
            
        # Сохраняем в состояние
        await context.update_data(max_id=max_id)
        settings = await db.get_settings()
        # Переходим к следующему шагу
        await context.set_state(AddDoctor.full_name)
        await event.message.answer(
            "✅ ID принят. Теперь введите ФИО доктора:\n" \
            "Пример: <i>Иванова Н.И.</i>")
        
    except ValueError:
        await event.message.answer("❌ Неверный формат. Введите числовой ID Telegram:", attachments=[kb.home_keyboard_inline()])
    except Exception as e:
        logger.error(f"Ошибка при обработке Telegram ID: {e}")
        await event.message.answer("⚠️ Произошла ошибка. Попробуйте ввести ID еще раз:", attachments=[kb.home_keyboard_inline()])

########################################################################## 
@router.message_created(AddDoctor.full_name)
async def process_add_full_name(event: MessageCreated, db: DataBase, context: MemoryContext):
    """Обработчик ввода ФИО доктора с поддержкой формата 'Фамилия И.О.'"""
    settings = await db.get_settings()
    try:
        full_name = event.message.body.text.strip()
        
        # Валидация ФИО
        if not full_name:
            await event.message.answer("❌ ФИО не может быть пустым. Введите еще раз:")
            return
            
        if len(full_name) < 2:
            await event.message.answer("❌ Слишком короткое ФИО. Минимум 2 символа. Введите еще раз:")
            return
            
        if len(full_name) > 100:
            await event.message.answer("❌ Слишком длинное ФИО. Максимум 100 символов. Введите еще раз:")
            return
            
        # Проверка на допустимые символы (буквы, пробелы, точки, дефисы)
        if not re.match(r'^[a-zA-Zа-яА-ЯёЁ\s\-\.]+$', full_name):
            await event.message.answer(
                "❌ ФИО может содержать только:\n"
                "- Буквы (А-Я, A-Z)\n"
                "- Пробелы\n"
                "- Точки (для инициалов)\n"
                "- Дефисы\n"
                "Пример: Иванов А.А. или Петрова-Смирнова М.С.\n"
                "Введите еще раз:"
            )
            return
            
        # Дополнительная проверка формата "Фамилия И.О."
        if re.search(r'[а-яА-ЯёЁ]\.[а-яА-ЯёЁ]\.?$', full_name):
            if len(full_name) < 5:  # Минимум "А.А."
                await event.message.answer("❌ Слишком короткие инициалы. Введите еще раз:")
                return
                
        # Сохраняем в состояние
        await context.update_data(full_name=full_name)
        
        # Переходим к следующему шагу
        await context.set_state(AddDoctor.phone)
        await event.message.answer(
            "✅ ФИО принято. Теперь введите телефон доктора в формате +79991234567:")
        
    except Exception as e:
        logger.error(f"Ошибка при обработке ФИО: {e}", exc_info=True)
        await event.message.answer(
            "⚠️ Произошла ошибка. Пожалуйста, введите ФИО еще раз:",
            attachments=[kb.admin_main_kb(bot_status=settings.bot_active, auto_schedule=settings.auto_schedule)]
        )
##########################################################################
@router.message_created(AddDoctor.phone)
async def process_add_phone(event: MessageCreated, context: MemoryContext, db: DataBase):
    """Обработчик ввода телефона с сохранением и уведомлением"""
    settings = await db.get_settings()
    user_id = event.chat.chat_id
    user_state = await db.get_state(user_id)
    try:
        phone = event.message.body.text.strip()
        
        # Валидация телефона
        if not re.match(r'^(\+7|8)\d{10}$', phone):
            await event.message.answer("❌ Неверный формат. Используйте +79991234567 или 89991234567")
            return
            
        # Приводим к формату +7
        formatted_phone = re.sub(r'^8', '+7', phone)
        
        # Получаем и дополняем данные
        data = await context.get_data()
        doctor_data = {
            'max_id': data.get('max_id'),
            'full_name': data['full_name'],
            'phone': formatted_phone
        }
        
        # Сохраняем в базу
        result = await db.add_doctor(doctor_data)
        
        if not result.get('success', False):
            await event.message.answer(f"❌ {result.get('message', 'Ошибка сохранения')}")
            return
            
        # Очищаем состояние
        await context.set_state(None)
        
        # Отправляем подтверждение админу
        admin_message = (
            "✅ Доктор добавлен:\n"
            f"ID: {result['doctor_id']}\n"
            f"ФИО: {doctor_data['full_name']}\n"
            f"Телефон: {formatted_phone}"
        )
        sent_mess = await event.message.answer(admin_message, attachments=[await kb.inline_back_admin()])
        # Обновляем состояние пользователя
        user_state.last_message_ids.append(sent_mess.message.body.mid)
        await db.update_state(user_state)
        # Пытаемся отправить уведомление доктору
        if doctor_data.get('max_id'):
            try:
                await event.bot.send_message(
                    chat_id=doctor_data['max_id'],
                    text=(
                        "🔔 Вас добавили в систему как доктора!\n"
                        f"Ваш ID: {result['doctor_id']}\n"
                        f"Статус: Активный"
                    )
                )
            except Exception:
                logger.warning(f"Доктор {result['doctor_id']} запретил сообщения")
            except Exception as e:
                logger.error(f"Ошибка уведомления доктора: {e}")
                
    except Exception as e:
        logger.error(f"Ошибка добавления доктора: {e}", exc_info=True)
        await event.message.answer(
            "⚠️ Произошла ошибка. Попробуйте позже.",attachments=[kb.admin_main_kb(
                    bot_status=settings.bot_active, auto_schedule=settings.auto_schedule
                )]
            
        )
##########################################################################



##############################################################################
# ========== ОБРАБОТЧИКИ ДЛЯ РЕГИСТРАТОРА ==========

@router.message_callback(F.callback.payload.startswith("accept_choice_doc_"))
async def accept_choice_doc(event: MessageCallback, db: DataBase, context: MemoryContext):
    """Регистратор выбирает врача для вызова"""
    try:
        call_id = int(event.callback.payload.split("_")[-1])
        await logger.info(f'accept_choice_doc: call_id={call_id}')
        
        # Получаем данные вызова
        call_data = await db.get_call_by_id(call_id)
        
        # Загружаем список врачей
        doctors = await db.get_active_doctors()
        
        if doctors:
            # Показываем выбор врача
            await event.message.edit(
                "👨‍⚕️ Выберите врача для этого вызова:",
                attachments=[await kb.choice_doctors(call_id, doctors)]
            )
        else:
            # Если врачей нет, просто подтверждаем вызов
            await logger.info(f'accept_choice_doc: врачей нет, подтверждаем вызов')
            
            # Удаляем старое сообщение с кнопками
            await event.message.delete()
            
            # Отправляем новое сообщение без кнопок с номером вызова
            call_number = call_data.call_number if call_data else call_id
            await event.message.answer(f"✅ Вызов #{call_number} принят (врач не назначен, так как список врачей пуст)")
            
            # Принимаем вызов
            await accept_call_without_doctor(event, db, call_id)
            return
        
    except Exception as e:
        await logger.error(f"Ошибка в accept_choice_doc: {e}")
        await event.message.answer("⚠️ Произошла ошибка", attachments=[kb.home_keyboard_inline()])

##############################################################################
@router.message_callback(F.callback.payload.startswith("accept_"))
async def accept_call_handler(event: MessageCallback, db: DataBase):
    """Обработчик принятия вызова с назначением врача"""
    try:
        parts = event.callback.payload.split("_")
        call_id = int(parts[1])
        doc_id = int(parts[2]) if len(parts) > 2 else None
        
        # Получаем данные вызова
        call_data = await db.get_call_by_id(call_id)
        if not call_data:
            await event.message.answer("❌ Вызов не найден")
            return
        
        # Обновляем статус вызова - передаем Enum
        await db.update_call_status(
            call_id=call_id,
            new_status='approved',
            rejection_reason=None,
            doc_id=doc_id
        )
        # Удаляем сообщение у регистратора
        await event.message.delete()
        
        # Отправляем подтверждение регистратору
        await event.message.answer(f"✅ Вызов #{call_data.call_number} принят.")
        
        # Уведомляем пациента
        try:
            await event.bot.send_message(
                chat_id=call_data.user_id,
                text=f"🏥 <b>Статус вызова #{call_data.call_number}</b>\n\n"
                     f"✅ Ваш вызов принят!\n"
                     f"👨‍⚕️ Врач приедет в течение дня.\n"
                     f"📞 Будьте на связи по телефону: {call_data.phone}",
                parse_mode="HTML",
                attachments=[await kb.main_menu_kb(call_data.user_id)]
            )
        except Exception as e:
            await logger.error(f"Ошибка отправки пациенту {call_data.user_id}: {e}")
        
        # Если назначен врач, отправляем ему уведомление
        if doc_id:
            doctor = await db.get_doctor(doc_id)
            if doctor and doctor.max_id:
                staff_message = (
                    "🚨 <b>Новый вызов</b>\n\n"
                    f"📋 <b>Номер вызова:</b> #{call_data.call_number}\n"
                    f"👤 <b>Пациент:</b> {call_data.full_name}\n"
                    f"🎂 <b>Дата рождения:</b> {call_data.birth_date}\n"
                    f"📞 <b>Телефон:</b> {call_data.phone}\n\n"
                    f"🏠 <b>Адрес:</b> {call_data.address}\n"
                    f"🚪 <b>Подъезд/этаж:</b> {call_data.address_details or 'не указано'}\n"
                    f"ℹ️ <b>Доп. информация:</b> {call_data.access_notes or 'нет'}\n\n"
                    f"🌡 <b>Температура:</b> {call_data.temperature or 'не указана'}\n"
                    f"🤒 <b>Симптомы:</b> {call_data.symptoms}\n"
                    f"🏥 <b>Больничный:</b> {'Нужен' if call_data.need_sick_leave else 'Не нужен'}"
                )
                try:
                    await event.bot.send_message(
                        chat_id=doctor.max_id,
                        text=staff_message
                    )
                    await logger.info(f"Вызов отправлен врачу {doctor.full_name}")
                except Exception as e:
                    await logger.error(f"Ошибка отправки врачу: {e}")
        
    except Exception as e:
        await logger.error(f"Ошибка в accept_call_handler: {e}")
        await event.message.answer("⚠️ Произошла ошибка", attachments=[kb.home_keyboard_inline()])

#===========================================================================
async def accept_call_without_doctor(event, db, call_id):
    """Принятие вызова без назначения врача"""
    try:
        await logger.info(f'accept_call_without_doctor: начало, call_id={call_id}')
        
        # Получаем данные вызова
        call_data = await db.get_call_by_id(call_id)
        if not call_data:
            await logger.error(f'accept_call_without_doctor: вызов {call_id} не найден')
            return
        
        await logger.info(f'accept_call_without_doctor: call_data получен, номер={call_data.call_number}')
        
        # Обновляем статус вызова - передаем Enum, НЕ строку и НЕ .value
        await db.update_call_status(
            call_id=call_id,
            new_status='approved', 
            rejection_reason=None,
            doc_id=None
        )
        await logger.info(f'accept_call_without_doctor: статус обновлен')
        
        # Удаляем предыдущие сообщения пациента
        user_state = await db.get_state(call_data.user_id)
        if user_state and user_state.last_message_ids:
            await db.delete_messages(user_state)
            user_state.last_message_ids = []
            await db.update_state(user_state)

        # Уведомляем пациента
        main_menu = kb.home_keyboard()
        sent_mess = await event.bot.send_message(
            chat_id=call_data.user_id,
            text=f"🏥 <b>Статус вызова #{call_data.call_number}</b>\n\n"
                 f"✅ Ваш вызов принят!\n"
                 f"👨‍⚕️ Врач приедет в течение дня.\n"
                 f"🏠 Главное меню => /Start"
        )
        # user_state.last_message_ids.append(sent_mess.message.body.mid)
        # await db.update_state(user_state)
        await logger.info(f'accept_call_without_doctor: вызов #{call_data.call_number} принят, пациент уведомлен')
        
    except Exception as e:
        await logger.error(f"Ошибка в accept_call_without_doctor: {e}")

#==========================================================================

@router.message_callback(F.callback.payload.startswith("reject_"))
async def reject_call_handler(event: MessageCallback, db: DataBase):
    """Обработчик отклонения вызова"""
    try:
        call_id = int(event.callback.payload.split("_")[1])
        
        # Получаем данные вызова
        call_data = await db.get_call_by_id(call_id)
        if not call_data:
            await event.message.answer("❌ Вызов не найден")
            return
        
        # Обновляем статус вызова - используем строку
        success = await db.update_call_status(
            call_id=call_id,
            new_status='rejected',  # строка, а не Enum
            rejection_reason="Отклонен call-центром",
            doc_id=None
        )
        
        if not success:
            await event.message.answer("❌ Ошибка при обновлении статуса")
            return
        
        # Удаляем сообщение регистратора с кнопками
        await event.message.delete()
        
        # Удаляем предыдущие сообщения пациента
        user_state = await db.get_state(call_data.user_id)
        if user_state and user_state.last_message_ids:
            await db.delete_messages(user_state)
            user_state.last_message_ids = []
            await db.update_state(user_state)
        
        # Отправляем уведомление пациенту
        main_menu = kb.home_keyboard()
        await event.bot.send_message(
            chat_id=call_data.user_id,
            text=f"🏥 Статус вызова #{call_data.call_number}\n\n"
                 f"❌ Ваш вызов отклонён.\n"
                 f"📝 Причина: Отклонен call-центром\n\n"
                 f"ℹ️ Для дополнительных вопросов обратитесь в регистратуру.",
            attachments=[main_menu]
        )
        
        # Отправляем подтверждение регистратору
        await event.message.answer(f"❌ Вызов #{call_data.call_number} отклонён")
        
        await logger.info(f'reject_call_handler: вызов #{call_data.call_number} отклонён')
        
    except Exception as e:
        await logger.error(f"Ошибка в reject_call_handler: {e}")
        await event.message.answer("⚠️ Произошла ошибка", attachments=[kb.home_keyboard_inline()])
###############################################################################
@router.message_created(DoctorCall.REASON)
async def process_reject_reason(event: MessageCreated, context: MemoryContext, db: DataBase):
    """Обработка причины отказа"""
    try:
        reason = event.message.body.text.strip()
        
        if len(reason) < 5:
            await send_error(event, context, db, "❌ Причина слишком короткая (минимум 5 символов):")
            return
        
        # Получаем ID вызова из контекста
        data = await context.get_data()
        call_id = data.get('reject_call_id')
        
        if not call_id:
            await event.message.answer("❌ Ошибка: вызов не найден")
            return
        
        # Получаем данные вызова
        call_data = await db.get_call_by_id(call_id)
        if not call_data:
            await event.message.answer("❌ Вызов не найден")
            return
        
        # Обновляем статус вызова
        await db.update_call_status(
            call_id=call_id,
            new_status='rejected',
            rejection_reason=reason,
            doc_id=None
        )
        
        # Удаляем сообщение пользователя с причиной
        await event.message.delete()
        
        # Отправляем подтверждение регистратору
        await event.message.answer(f"❌ Вызов #{call_data.call_number} отклонён. Причина сохранена.")
        
        # Уведомляем пациента
        try:
            await event.bot.send_message(
                chat_id=call_data.user_id,
                text=f"🏥 <b>Статус вызова #{call_data.call_number}</b>\n\n"
                     f"❌ Ваш вызов отклонён.\n"
                     f"📝 <b>Причина:</b> {reason}\n\n"
                     f"ℹ️ Для дополнительных вопросов обратитесь в регистратуру.",
                parse_mode="HTML",
                attachments=[await kb.main_menu_kb(call_data.user_id)]
            )
            await logger.info(f'process_reject_reason: уведомление отправлено пациенту {call_data.user_id}')
        except Exception as e:
            await logger.error(f"Ошибка отправки пациенту: {e}")
        
        # Очищаем состояние
        await context.set_state(None)
        await context.update_data(reject_call_id=None)
        
        # Удаляем сообщение с кнопками (если нужно)
        user_state = await db.get_state(event.chat.chat_id)
        if user_state and user_state.last_message_ids:
            await db.delete_messages(user_state)
            user_state.last_message_ids = []
            await db.update_state(user_state)
        
        await logger.info(f'process_reject_reason: вызов #{call_data.call_number} отклонен')
        
    except Exception as e:
        await logger.error(f"Ошибка в process_reject_reason: {e}")
        await event.message.answer("⚠️ Произошла ошибка", attachments=[kb.home_keyboard_inline()])
#==============================================================================
async def send_error(event, context, db, text):
    """Отправка ошибки"""
    await logger.info(f'send_error: {text}')
    sent_mess = await event.message.answer(text, attachments=[kb.cancel_kb()])
    user_state = await db.get_state(event.chat.chat_id)
    if user_state:
        user_state.last_message_ids.append(sent_mess.message.body.mid)
        await db.update_state(user_state)
################################################################################

@router.message_callback(F.callback.payload.startswith("stats_day_"))
async def handle_day_selection(event: MessageCallback, db: DataBase):
    """Обработчик выбора даты для статистики"""
    try:
        parts = event.callback.payload.split("_")
        year = int(parts[2])
        month = int(parts[3])
        day = int(parts[4])
        
        user_id = event.chat.chat_id
        user_state = await db.get_state(user_id)
        
        # Удаляем предыдущие сообщения если нужно
        if user_state and user_state.last_message_ids:
            await db.delete_messages(user_state)
            user_state.last_message_ids = []
            await db.update_state(user_state)

        await logger.info(f'handle_day_selection: user={user_id}, date={year}-{month}-{day}')

        # Форматируем дату
        date_obj = datetime(year, month, day)
        formatted_date = date_obj.strftime("%Y-%m-%d")
        display_date = date_obj.strftime("%d-%m-%Y")
        
        # Получаем статистику за выбранный день
        stats = await db.get_daily_statistics(formatted_date)
        
        if not stats:
            await event.message.answer("⚠️ Ошибка при загрузке статистики")
            return
        
        # Форматируем сообщение
        message = (
            f"📊 <b>Статистика на {display_date}</b>\n\n"
            f"• Всего вызовов: <b>{stats['total_calls']}</b>\n"
            f"• ✅ Принято: <b>{stats['approved_calls']}</b>\n"
            f"• ❌ Отклонено: <b>{stats['rejected_calls']}</b>\n"
            f"• ⏳ В ожидании: <b>{stats['pending_calls']}</b>\n"
            f"• 🚫 Отменено: <b>{stats['cancelled_calls']}</b>"
        )
        
        # Отправляем сообщение с клавиатурой
        sent_msg = await event.message.answer(
            message,
            attachments=[kb.admin_kb()]
        )
        
        # Сохраняем ID сообщения
        user_state = await db.get_state(user_id)
        user_state.last_message_ids.append(sent_msg.message.body.mid)
        await db.update_state(user_state)
        
    except Exception as e:
        await logger.error(f"Ошибка при загрузке статистики: {e}")
        await event.message.answer("⚠️ Ошибка при загрузке статистики", attachments=[kb.home_keyboard_inline()])

#############################################################################
@router.message_callback(F.callback.payload.startswith("period_custom"))
async def handle_period_custom(event: MessageCallback, context: MemoryContext, db: DataBase):
    """Начинаем выбор периода - сначала начальную дату"""
    await logger.info(f'ID_MAX:{event.chat.chat_id}|=>handle_period_custom')
    
    # Сбрасываем предыдущий выбор в FSM
    await context.update_data(
        stats_period_start=None,
        stats_period_end=None
    )
    try:
        # Показываем календарь для выбора начальной даты
        await event.message.edit(
            "📅 Выберите **начальную** дату периода:",
            attachments=[await kb.stats_calendar_kb(select_start=True)]
        )
        
    except Exception as e:
        await logger.error(f"Ошибка при выборе начальной даты: {e}")
        await event.message.answer("⚠️ Ошибка при выборе начальной даты", attachments=[kb.home_keyboard_inline()])
############################################################################

@router.message_callback(F.callback.payload.startswith("stats_calendar_"))
async def handle_stats_calendar_nav(event: MessageCallback, db: DataBase):
    """Навигация по месяцам в календаре"""
    # Данные: stats_calendar_2024_5_start
    parts = event.callback.payload.split("_")
    year = int(parts[2])
    month = int(parts[3])
    select_start = parts[4] == "start"  # start или end
    print("stats_calendar_")
    print(f"parts:{parts}")      # "📅 Выберите **начальную** дату периода:

    try:
        await event.message.edit(
            attachments=[await kb.stats_calendar_kb(
                select_start=select_start,
                year=year,
                month=month
            )]
        )
    except Exception as e:
        await logger.error(f"Ошибка stats_calendar_ {e}")
        await event.message.answer("⚠️ Ошибка при выборе начальной даты", attachments=[kb.home_keyboard_inline()])
############################################################################

@router.message_callback(F.callback.payload.startswith("stats_select_date_"))
async def handle_stats_select_date(event: MessageCallback, context: MemoryContext, db: DataBase):
    """Выбор конкретной даты в календаре периода"""
    # Данные: stats_select_date_2024_5_15_start
    parts = event.callback.payload.split("_")
    year = int(parts[3])
    month = int(parts[4])
    day = int(parts[5])
    is_start_date = parts[6] == "start"  # start или end
    
    # Форматируем дату
    date_str = f"{day:02d}.{month:02d}.{year}"
    await logger.info(f"handle_stats_select_date: выбрана {date_str}, is_start={is_start_date}")
    
    # Получаем текущие данные
    user_data = await context.get_data()
    
    # Сохраняем дату
    if is_start_date:
        await context.update_data(stats_period_start=date_str)
        await logger.info(f"Сохранена начальная дата: {date_str}")
        
        # Показываем календарь для выбора конечной даты
        await event.message.edit(
            "📅 Выберите **конечную** дату периода:",
            attachments=[await kb.stats_calendar_kb(select_start=False, year=year, month=month)]
        )
    else:
        await context.update_data(stats_period_end=date_str)
        await logger.info(f"Сохранена конечная дата: {date_str}")
        
        # Получаем обе даты
        data = await context.get_data()
        start_date = data.get('stats_period_start')
        end_date = data.get('stats_period_end')
        
        if not start_date or not end_date:
            await event.message.answer("❌ Ошибка: не выбрана начальная дата")
            return
        
        # Проверяем корректность периода
        from datetime import datetime as dt
        start = dt.strptime(start_date, "%d.%m.%Y")
        end = dt.strptime(end_date, "%d.%m.%Y")
        
        if start > end:
            await event.message.answer("❌ Начальная дата не может быть позже конечной! Поменяйте местами.")
            # Меняем местами
            await context.update_data(
                stats_period_start=end_date,
                stats_period_end=start_date
            )
            start_date, end_date = end_date, start_date
        
        # Показываем кнопку подтверждения
        builder = InlineKeyboardBuilder()
        builder.row(
            CallbackButton(
                text=f"📊 Скачать Excel за период",
                payload=f"export_period_{start_date}_{end_date}"
            )
        )
        builder.row(
            CallbackButton(text="🔄 Выбрать другие даты", payload="period_custom")
        )
        builder.row(
            CallbackButton(text="⚙️ В админ панель", payload="admin")
        )
        
        await event.message.edit(
            f"📊 **Выбранный период:**\n"
            f"• Начало: {start_date}\n"
            f"• Конец: {end_date}\n\n"
            f"Нажмите кнопку для экспорта:",
            attachments=[builder.as_markup()]
        )
    
    # В MAXAPI нет answer(), просто логируем
    await logger.info(f"Выбрана {'начальная' if is_start_date else 'конечная'} дата: {date_str}")
############################################################################

##############################################################################

@router.message_callback(F.callback.payload.startswith("export_period_"))
async def handle_export_period(event: MessageCallback, db: DataBase):
    """Экспорт Excel за выбранный период"""
    user_id = event.chat.chat_id
    user_state = await db.get_state(user_id)
    
    if user_state and user_state.last_message_ids:
        await db.delete_messages(user_state)
        user_state.last_message_ids = []
        await db.update_state(user_state)
    
    try:
        # Данные: export_period_01.03.2026_30.03.2026
        parts = event.callback.payload.split("_")
        start_date_with_dots = parts[2]  # 01.03.2026
        end_date_with_dots = parts[3]    # 30.03.2026
        print(f"print:{parts}")


        # Конвертируем для БД: 01.03.2026 -> 01-03-2026
        start_date_for_db = start_date_with_dots.replace('.', '-')
        end_date_for_db = end_date_with_dots.replace('.', '-')

        await logger.info(f'Экспорт за период: {start_date_for_db} - {end_date_for_db}')

        # Получаем анкеты за период (метод ожидает формат ДД-ММ-ГГГГ)
        questionnaires = await db.get_questionnaires_by_period(start_date_for_db, end_date_for_db)
        
        if not questionnaires:
            sent_msg = await event.message.answer(
                f"📭 Нет данных за период {start_date_with_dots} - {end_date_with_dots}",
                attachments=[kb.admin_kb()]
            )
            user_state.last_message_ids.append(sent_msg.message.body.mid)
            await db.update_state(user_state)
            return
        
        excel_file_data = await generate_doctor_calls_excel(questionnaires, db)
        
        # Сохраняем в папку temp
        import os
        temp_filename = f"temp/export_{start_date_with_dots}_{end_date_with_dots}_{user_id}.xlsx"
        
        with open(temp_filename, 'wb') as f:
            f.write(excel_file_data)
        
        # Отправляем файл
        file_attachment = InputMedia(path=temp_filename)
        
        sent_msg = await event.bot.send_message(
            chat_id=user_id,
            text=f"📊 Вызовы врачей за период:\n{start_date_with_dots} - {end_date_with_dots}\n\nНайдено вызовов: {len(questionnaires)}",
            attachments=[file_attachment, kb.admin_kb()]
        )
        
        # Удаляем временный файл
        try:
            os.remove(temp_filename)
        except:
            pass
        
        user_state.last_message_ids.append(sent_msg.message.body.mid)
        await db.update_state(user_state)
        await logger.info(f"Файл отправлен пользователю {user_id}")

    except Exception as e:
        await logger.error(f"Критическая ошибка экспорта: {e}")
        sent_msg = await event.message.answer(
            f"⚠️ Ошибка при экспорте: {str(e)[:200]}",
            attachments=[kb.admin_kb()]
        )
        user_state.last_message_ids.append(sent_msg.message.body.mid)
        await db.update_state(user_state)

##############################################################################

# @router.message_callback(F.callback.payload.startswith("export_"))
# async def handle_export(event: MessageCallback, db: DataBase):
#     """Обработчик экспорта: генерирует и отправляет Excel-файл."""
#     user_id = event.chat.chat_id
#     user_state = await db.get_state(user_id)
    
#     if user_state and user_state.last_message_ids:
#         await db.delete_messages(user_state)
#         user_state.last_message_ids = []
#         await db.update_state(user_state)

#     try:
#         date_str = event.callback.payload.split("_")[1]
#         await logger.info(f'Экспорт данных за {date_str} для пользователя {user_id}')
        
#         questionnaires = await db.get_questionnaires_by_date(date_str)
#         if not questionnaires:
#             sent_msg = await event.message.answer(
#                 f"📭 Нет данных для экспорта за {date_str}.",
#                 attachments=[kb.admin_kb()]
#             )
#             user_state.last_message_ids.append(sent_msg.message.body.mid)
#             await db.update_state(user_state)
#             return

#         # Генерация Excel-файла
#         excel_file_data = await generate_doctor_calls_excel(questionnaires, db)
        
#          # Создаем папку temp если её нет
#         temp_dir = "temp"
#         if not os.path.exists(temp_dir):
#             os.makedirs(temp_dir)
        
#         # Сохраняем временный файл в папку temp
#         temp_filename = os.path.join(temp_dir, f"export_{date_str}_{user_id}.xlsx")
#         with open(temp_filename, 'wb') as f:
#             f.write(excel_file_data)
        
#         await logger.info(f'Временный файл создан: {temp_filename}')
        
#         # Отправляем файл
#         file_attachment = InputMedia(
#             path=temp_filename
#         )
        
#         sent_msg = await event.bot.send_message(
#             chat_id=user_id,
#             text=f"📊 Выгрузка данных за {date_str} выполнена.\nНайдено вызовов: {len(questionnaires)}.",
#             attachments=[file_attachment]
#         )
        
#         # Удаляем временный файл
#         try:
#             os.unlink(temp_filename)
#             await logger.info(f'Временный файл удален: {temp_filename}')
#         except Exception as del_err:
#             await logger.error(f'Ошибка удаления временного файла: {del_err}')
        
#         user_state.last_message_ids.append(sent_msg.message.body.mid)
#         await db.update_state(user_state)
#         await logger.info(f"Файл отправлен пользователю {user_id}")

#     except Exception as e:
#         await logger.error(f"Критическая ошибка экспорта: {e}")
#         sent_msg = await event.message.answer(
#             f"⚠️ Ошибка при экспорте: {str(e)[:200]}",
#             attachments=[kb.home_keyboard_inline()]
#         )
#         user_state.last_message_ids.append(sent_msg.message.body.mid)
#         await db.update_state(user_state)

################################################################################
async def generate_doctor_calls_excel(questionnaires: list, db) -> bytes:
    """Генерирует Excel-файл из списка вызовов и возвращает его в виде байтов."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Вызовы врачей"
    
    # --- Настройки для печати ---
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # --- Заголовки ---
    headers = ["№", "ФИО", "ДР", "Телефон", "Адрес", "Под/Эт", "Темп.", "Симптомы", 
               "Б/л", "Время вызова", "Статус", "Примечания", "Врач"]
    ws.append(headers)
    
    # --- Стили ---
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True, size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    wrap_text_alignment = Alignment(wrap_text=True, vertical='top')
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # --- Словарь статусов (ключи - строки) ---
    status_display = {
        'new': '🟡 Новый',
        'approved': '🟢 Принят',
        'rejected': '🔴 Отклонен',
        'cancelled': '⚪ Отменен',
        'pending_cancellation': '🟠 Ожидает отмены'
    }
    
    # --- Заполнение данными ---
    for row_num, call in enumerate(questionnaires, start=2):
        # Получаем имя врача
        doctor_name = ""
        if call.doctor_id:
            doctor = await db.get_doctor(call.doctor_id)
            doctor_name = doctor.full_name if doctor else ""
        
        # Преобразуем статус в строку (на всякий случай)
        status = str(call.status) if call.status else 'new'
        status_text = status_display.get(status, status)
        
        # Обрезаем симптомы
        symptoms = str(call.symptoms) if call.symptoms else ""
        if len(symptoms) > 100:
            symptoms = symptoms[:97] + '...'
        
        # Форматируем дату создания
        created_at_str = ""
        if call.created_at:
            created_at_str = call.created_at.strftime('%d.%m.%Y %H:%M')
        
        ws.append([
            str(call.call_number) if call.call_number else str(row_num - 1),
            str(call.full_name) if call.full_name else '',
            str(call.birth_date) if call.birth_date else '',
            str(call.phone) if call.phone else '',
            str(call.address) if call.address else '',
            str(call.address_details) if call.address_details else 'не указано',
            str(call.temperature) if call.temperature else 'не указана',
            symptoms,
            '✓' if call.need_sick_leave else '',
            created_at_str,
            status_text,
            str(call.access_notes) if call.access_notes else '',
            doctor_name
        ])
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            cell.alignment = wrap_text_alignment
    
    # --- Настройка ширины столбцов ---
    column_widths = {'A':6, 'B':25, 'C':12, 'D':15, 'E':30, 'F':12, 'G':10, 
                     'H':40, 'I':6, 'J':16, 'K':15, 'L':30, 'M':25}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream.getvalue()

##############################################################################