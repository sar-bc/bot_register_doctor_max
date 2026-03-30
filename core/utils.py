from typing import TYPE_CHECKING, Optional, List
from aiogram.types import Message
from database.Database import DataBase
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from datetime import datetime
from functools import wraps
from aiogram import types
from typing import Union, Callable
from core.dictionary import *
from database.Database import DataBase
import app.keyboards as kb
from datetime import datetime, timedelta, time




if TYPE_CHECKING:
    from aiogram import Bot
    from aiogram.fsm.context import FSMContext
    from database.models import UserState

logger = logging.getLogger(__name__)
####################################################################
def format_minutes(minutes: int) -> str:
    """Форматирует правильное склонение слова 'минута'"""
    if minutes == 0:
        return "менее минуты"
    
    last_digit = minutes % 10
    last_two_digits = minutes % 100
    
    if last_digit == 1 and last_two_digits != 11:
        return f"{minutes} минута"
    elif last_digit in [2, 3, 4] and last_two_digits not in [12, 13, 14]:
        return f"{minutes} минуты"
    else:
        return f"{minutes} минут"
####################################################################
dop_minutes = 10
def check_bot_status_universal(func: Callable):
    @wraps(func)
    async def wrapper(update: Union[types.Message, types.CallbackQuery], *args, **kwargs):
        # Определяем тип обновления
        if isinstance(update, types.CallbackQuery):
            message = update.message
            user_id = update.from_user.id
            is_callback = True
        else:
            message = update
            user_id = update.from_user.id
            is_callback = False
        db = DataBase()
        # Проверяем статус бота
        current_time = datetime.now()
        settings = await db.get_settings()
        status_bot = await db.get_bot_status()
        new_time = settings.last_changed+timedelta(minutes=dop_minutes)
        # logger.info(f'settings.last_changed:{settings.last_changed}')
        # logger.info(f'settings.last_changed+10:{settings.last_changed+timedelta(minutes=10)}')
        # logger.info(f'status_bot:{status_bot}')

        if not status_bot and current_time > new_time:
            user_state = await db.get_state(user_id)
            await db.delete_messages(user_state)

            sent_mess = await message.answer(
                CALLS_FINISHED_TODAY_TEXT % (
                    settings.weekday_start.strftime('%H:%M'),
                    settings.weekday_end.strftime('%H:%M'),
                    settings.weekend_start.strftime('%H:%M'),
                    settings.weekend_end.strftime('%H:%M')
                ),
                reply_markup=kb.home_keyboard()
            )
            user_state.last_message_ids.append(sent_mess.message_id)
            await db.update_state(user_state)
            
            if is_callback:
                await update.answer()  # Ответ на колбэк
            return
        elif not status_bot and current_time < new_time:
            time_until_close = new_time - current_time
            await update.answer(
                    f"⏳ До закрытия записи осталось {format_minutes(int(time_until_close.total_seconds() // 60))}!"
                )
        return await func(update, *args, **kwargs)
    return wrapper
####################################################################
async def generate_doctor_calls_excel(questionnaires: list, db) -> bytes:
    """Генерация Excel файла, оптимизированного для печати"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Вызовы врачей"
    
    # Настройки страницы для печати
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # Простые колонтитулы без форматирования (чтобы избежать ошибки)
    ws.oddHeader.center.text = "Вызовы врачей на " + datetime.now().strftime('%d.%m.%Y')
    
    # Заголовки столбцов
    headers = [
        "№", "ФИО", "ДР", "Телефон",
        "Адрес", "Под/Эт", "Темп.",
        "Симптомы", "Б/л", "Время вызова",
        "Статус", "Примечания", "Врач"  # Добавлен новый столбец
    ]
    ws.append(headers)
    
    # Стили для оформления
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True, size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    wrap_text = Alignment(wrap_text=True, vertical='top')
    
    # Форматируем заголовки
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Заполняем данными
    for idx, questionnaire in enumerate(questionnaires, start=1):
        symptoms = questionnaire.symptoms
        if len(symptoms) > 100:
            symptoms = symptoms[:97] + '...'
        
        # Получаем данные врача
        doctor_name = ""
        if questionnaire.doctor_id:
            doc = await db.get_doctor(questionnaire.doctor_id)
            doctor_name = doc.full_name if doc else ""
        
        row = [
            questionnaire.call_number, #idx,
            questionnaire.full_name,
            questionnaire.birth_date,
            questionnaire.phone,
            questionnaire.address,
            getattr(questionnaire, 'address_details', 'не указано'),
            getattr(questionnaire, 'temperature', 'не указана'),
            symptoms,
            '✓' if getattr(questionnaire, 'need_sick_leave', False) else '',
            questionnaire.created_at.strftime('%d.%m.%Y %H:%M'),
            db.STATUS_DISPLAY.get(questionnaire.status, 'Неизвестен'),
            getattr(questionnaire, 'access_notes', ''),
            doctor_name  # Добавляем ФИО врача
        ]
        ws.append(row)
        
        # Форматируем строку с данными
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=idx+1, column=col)
            cell.border = border
            cell.alignment = wrap_text
    
    # Настраиваем ширину столбцов
    column_widths = {
        'A': 5,   # №
        'B': 25,  # ФИО
        'C': 12,  # Дата рождения
        'D': 15,  # Телефон
        'E': 30,  # Адрес
        'F': 15,  # Подъезд/этаж
        'G': 10,  # Температура
        'H': 40,  # Симптомы
        'I': 5,   # Больничный
        'J': 16,  # Дата/время
        'K': 15,  # Статус
        'L': 30,  # Примечания
        'M': 25   # Врач (новый столбец)
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Настраиваем высоту строк
    for row in range(2, len(questionnaires) + 2):
        ws.row_dimensions[row].height = 30
    
    # Добавляем автофильтр и закрепляем заголовки
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'
    
    # Сохраняем в бинарный поток
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file.getvalue()

####################################################################
class MessageCleaner:
    @staticmethod
    async def delete_inline_messages(bot: 'Bot', user_state: 'UserState') -> None:
        """Безопасное удаление inline-сообщений"""
        if not user_state.last_message_ids:
            return

        remaining_ids = []
        for msg_id in user_state.last_message_ids:
            try:
                # Проверяем, является ли сообщение inline
                await bot.edit_message_reply_markup(
                    chat_id=user_state.user_id,
                    message_id=msg_id,
                    reply_markup=None
                )
                await bot.delete_message(
                    chat_id=user_state.user_id,
                    message_id=msg_id
                )
                logger.debug(f"Удалено inline-сообщение {msg_id}")
            except Exception as e:
                logger.warning(f"Не удалось удалить сообщение {msg_id}: {str(e)}")
                remaining_ids.append(msg_id)

        user_state.last_message_ids = remaining_ids

    @staticmethod
    async def cleanup_user_messages(bot: 'Bot', user_state: 'UserState') -> None:
        """Полная очистка всех сообщений"""
        if not user_state.last_message_ids:
            return

        for msg_id in user_state.last_message_ids:
            try:
                await bot.delete_message(
                    chat_id=user_state.user_id,
                    message_id=msg_id
                )
                logger.debug(f"Удалено сообщение {msg_id}")
            except Exception as e:
                logger.warning(f"Не удалось удалить сообщение {msg_id}: {str(e)}")

        user_state.last_message_ids = []

async def reset_user_state(user_id: int, db: DataBase, state: 'FSMContext') -> None:
    """
    Полный сброс состояния пользователя
    Args:
        user_id: ID пользователя в Telegram
        db: Экземпляр базы данных
        state: Контекст FSM
    """
    try:
        # Очищаем состояние FSM
        await state.clear()
        
        # Получаем и очищаем состояние пользователя
        user_state = await db.get_state(user_id)
        await MessageCleaner.cleanup_user_messages(state.bot, user_state)
        
        # Сбрасываем историю сообщений
        user_state.last_message_ids = []
        await db.update_state(user_state)
        
        logger.info(f"Сброшено состояние для пользователя {user_id}")
    except Exception as e:
        logger.error(f"Ошибка сброса состояния: {str(e)}")
        raise

async def send_managed_message(
    message: Message,
    db: DataBase,
    text: str,
    reply_markup=None,
    cleanup_existing: bool = True
) -> Message:
    """
    Отправляет сообщение с управлением историей
    
    Args:
        message: Объект сообщения Aiogram
        db: Экземпляр базы данных
        text: Текст сообщения
        reply_markup: Клавиатура (опционально)
        cleanup_existing: Нужно ли удалять предыдущие сообщения
        
    Returns:
        Отправленное сообщение
    """
    try:
        user_state = await db.get_state(message.from_user.id)
        
        if cleanup_existing:
            await MessageCleaner.delete_inline_messages(message.bot, user_state)
        
        # Отправляем новое сообщение
        sent_msg = await message.answer(
            text=text,
            reply_markup=reply_markup
        )
        
        # Обновляем состояние
        user_state.last_message_ids.append(sent_msg.message_id)
        await db.update_state(user_state)
        
        logger.debug(f"Отправлено управляемое сообщение {sent_msg.message_id}")
        return sent_msg
        
    except Exception as e:
        logger.error(f"Ошибка отправки управляемого сообщения: {str(e)}")
        raise